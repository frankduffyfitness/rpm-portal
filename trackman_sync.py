#!/usr/bin/env python3
"""
Trackman master sheet → trackman_portal.json

Reads the RPM Trackman Master Sheet and produces trackman_portal.json — the velo
equivalent of forcedecks_portal.json. Accepts either:

  * the .xlsx workbook directly (reads the "Master Trackman Data" and
    "Athlete Program Links" tabs), or
  * a CSV export of the Master Trackman Data tab (+ optional program-links CSV).

Usage:
    python3 trackman_sync.py "RPM Trackman Master Sheet.xlsx" > trackman_portal.json
    python3 trackman_sync.py master.csv [program_links.csv]   > trackman_portal.json

The output JSON shape mirrors forcedecks_portal.json closely so generate_portal_data.py
can treat it the same way.
"""
import csv
import json
import os
import re
import sys
from datetime import datetime, date, timezone
from collections import defaultdict


# ─── Config ──────────────────────────────────────────────────────────────────

# Aliases the master sheet uses → real athlete names.
NAME_ALIASES = {
    "GLV":           "Gavin Laya-Vetell",
    "IRP":           "Isaiah Rubin-Patel",
    "Bob Billiams":  "Rob Williams",
}

# Session types that are intentionally submax and should NOT count toward
# "best ever" peak/avg leaderboards. Latest-session display still includes them.
SUBMAX_SESSION_TYPES = {
    "Low Effort",
    "Rehab",
}

# Notes patterns that flag a session as having Trackman data quality issues.
# Flagged sessions are excluded from the leaderboard / sparkline; latest-session
# still shows them but with a small flag in the data so the UI can mark them.
TRACKMAN_ISSUE_PATTERNS = [
    r"trackman\s+(issues?|errors?|not\s+(?:really\s+)?working)",
    r"few\s+pitches",
    r"only\s+\d+\s+pitches?",
    r"iffy\s+calibration",
    r"trackman\s+(?:sort\s+of\s+)?in\s+and\s+out",
]
_ISSUE_RE = re.compile("|".join(TRACKMAN_ISSUE_PATTERNS), re.IGNORECASE)

# Column header synonyms — the sheet has used both over time.
COL_PEAK = ("Peak FB", "Peak FB Velo")
COL_AVG  = ("Avg FB", "Avg FB Velo")
COL_NAME = ("Athlete Name", "Athlete")
COL_PROG = ("Athlete Program URL", "Program URL")


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _cell(row, *names):
    """Flexible (case-insensitive) column getter for a header→value dict."""
    for n in names:
        if n in row:
            return row[n]
    low = {str(k).strip().lower(): v for k, v in row.items()}
    for n in names:
        if n.lower() in low:
            return low[n.lower()]
    return None


def parse_date(s):
    """Accept a datetime/date object (from .xlsx) or a string (M/D/YYYY or ISO).
    Return ISO YYYY-MM-DD or None."""
    if isinstance(s, (datetime, date)):
        return s.strftime("%Y-%m-%d")
    s = (s if isinstance(s, str) else (str(s) if s is not None else "")).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except (ValueError, TypeError):
            continue
    return None


def parse_velo(s):
    """Parse a velo value. Accepts numbers (from .xlsx) or strings, handling
    typos like '72/4' → 72.4. Returns None if missing or out of sane bounds."""
    if isinstance(s, (int, float)):
        v = float(s)
    else:
        s = (s or "").strip()
        if not s:
            return None
        # Common typo: forward slash where decimal should be (e.g. "72/4")
        if re.match(r"^\d+/\d$", s):
            s = s.replace("/", ".")
        try:
            v = float(s)
        except ValueError:
            return None
    # Sanity bounds — peak FB velo for any human pitcher is 30-110 mph
    if v < 30 or v > 110:
        return None
    return round(v, 1)


def normalize_name(raw):
    """Trim whitespace, apply aliases."""
    n = (raw if isinstance(raw, str) else (str(raw) if raw is not None else "")).strip()
    return NAME_ALIASES.get(n, n)


def is_flagged(notes):
    """Sessions with Trackman data quality issues per the Notes column."""
    if not notes:
        return False
    return bool(_ISSUE_RE.search(notes))


# ─── Row parsing (shared by CSV and XLSX) ────────────────────────────────────

def parse_master_rows(rows):
    """Yield normalized session dicts from an iterable of header→value dicts."""
    for row in rows:
        date_iso = parse_date(_cell(row, "Date"))
        name = normalize_name(_cell(row, *COL_NAME))
        peak = parse_velo(_cell(row, *COL_PEAK))
        avg  = parse_velo(_cell(row, *COL_AVG))
        stype = (str(_cell(row, "Session Type") or "").strip()) or "Other"
        notes = str(_cell(row, "Notes") or "").strip()
        program_url = str(_cell(row, *COL_PROG) or "").strip()

        # Reject rows missing the essentials
        if not date_iso or not name or peak is None:
            continue

        yield {
            "date": date_iso,
            "name": name,
            "peakVelo": peak,
            "avgVelo": avg,
            "sessionType": stype,
            "notes": notes,
            "programUrl": program_url,
            "isFlagged": is_flagged(notes),
            "isSubmax": stype in SUBMAX_SESSION_TYPES,
        }


def _xlsx_rows(path, *preferred_sheets):
    """Yield header→value dicts from a worksheet (first preferred match, else first sheet)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = None
    for name in preferred_sheets:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        ws = wb.worksheets[0]
    headers = None
    for r in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(c).strip() if c is not None else "" for c in r]
            continue
        yield {headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))}


# ─── Loaders ─────────────────────────────────────────────────────────────────

def load_master(csv_path):
    """Yield normalized session dicts from the Master Trackman Data CSV."""
    with open(csv_path, newline="", encoding="utf-8") as f:
        yield from parse_master_rows(csv.DictReader(f))


def load_master_xlsx(path):
    """Yield normalized session dicts from the Master Trackman Data tab of an .xlsx."""
    yield from parse_master_rows(_xlsx_rows(path, "Master Trackman Data"))


def load_program_links(csv_path):
    """Optional second-tab CSV: Athlete Name → Program URL."""
    if not csv_path:
        return {}
    out = {}
    with open(csv_path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            name = normalize_name(_cell(row, *COL_NAME))
            url = str(_cell(row, "Program URL", "Athlete Program URL") or "").strip()
            if name and url:
                out[name] = url
    return out


def load_program_links_xlsx(path):
    """Athlete Program Links tab of an .xlsx: Athlete Name → Program URL."""
    out = {}
    for row in _xlsx_rows(path, "Athlete Program Links"):
        name = normalize_name(_cell(row, *COL_NAME))
        url = str(_cell(row, "Program URL", "Athlete Program URL") or "").strip()
        if name and url:
            out[name] = url
    return out


def build_portal(sessions, program_links):
    """Group by athlete, sort sessions newest-first, build the portal JSON."""
    by_athlete = defaultdict(list)
    for s in sessions:
        by_athlete[s["name"]].append(s)

    athletes = {}
    for name, rows in by_athlete.items():
        rows.sort(key=lambda s: s["date"], reverse=True)
        program_url = program_links.get(name) or rows[0].get("programUrl") or ""
        athletes[name] = {
            "name": name,
            "programUrl": program_url,
            "sessions": [
                {
                    "date":        s["date"],
                    "peakVelo":    s["peakVelo"],
                    "avgVelo":     s["avgVelo"],
                    "sessionType": s["sessionType"],
                    "notes":       s["notes"],
                    "isFlagged":   s["isFlagged"],
                    "isSubmax":    s["isSubmax"],
                    **({"provisional": True} if s.get("provisional") else {}),
                }
                for s in rows
            ],
        }
    merge_supplement(athletes)
    merge_report_sessions(athletes)
    return {
        "lastSyncedAt": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "athletes": athletes,
    }


def merge_supplement(athletes, path="master_sheet_supplement.csv"):
    """Official session rows Frank pastes directly (same columns as the master
    sheet). Applied AFTER the sheet: rows already present in the sheet are
    skipped, so once a future sheet export contains them this file is inert.
    These are official (coach-confirmed) rows, so they take precedence over
    the provisional PDF fill that runs after this."""
    if not os.path.exists(path):
        return
    added = 0
    with open(path, newline="", encoding="utf-8") as f:
        for s in parse_master_rows(csv.DictReader(f)):
            name = s["name"]
            if name not in athletes:
                athletes[name] = {"name": name, "programUrl": "", "sessions": []}
            ath = athletes[name]
            if any(x["date"] == s["date"] for x in ath["sessions"]):
                continue
            s.pop("programUrl", None)
            ath["sessions"].append(s)
            added += 1
        for ath in athletes.values():
            ath["sessions"].sort(key=lambda x: x["date"], reverse=True)
    if added:
        print(f"[trackman_sync] added {added} official session(s) from supplement", file=sys.stderr)


def merge_report_sessions(athletes, reports_path="trackman_reports.json"):
    """Fill master-sheet gaps from the bullpen-report PDFs.

    The coach's master sheet typically lags the PDF exports by a few days, so a
    pitcher's newest session (and any velo PR in it) was invisible to the velo
    tracker even though it showed on his Bullpen Breakdown. For any (athlete,
    date) present in trackman_reports.json but missing from the master sheet,
    append a PROVISIONAL session using the report's fastball numbers. These
    count toward peak velo and appear in history (labeled), but are excluded
    from average/trend math until the coach's official row arrives with the
    session type (which then replaces them on the next sync)."""
    if not os.path.exists(reports_path):
        return
    try:
        rep = json.load(open(reports_path))
    except Exception:
        return
    added = 0
    for name, r in (rep.get("athletes") or {}).items():
        ath = athletes.get(name)
        if not ath:
            continue  # only fill gaps for athletes the coach already tracks
        have = {s["date"] for s in ath["sessions"]}
        for s in r.get("sessions", []):
            if s["date"] in have:
                continue
            fb = next((t for t in s.get("types", []) if t.get("name") == "Fastball"), None)
            peak = (fb or {}).get("veloMax") or max(
                (t.get("veloMax") or 0 for t in s.get("types", [])), default=0)
            if not peak:
                continue
            ath["sessions"].append({
                "date": s["date"],
                "peakVelo": round(peak, 1),
                "avgVelo": round(fb["veloAvg"], 1) if fb and fb.get("veloAvg") else None,
                "sessionType": "Bullpen",
                "notes": "Auto-added from TrackMan report",
                "isFlagged": False,
                "isSubmax": False,
                "provisional": True,
            })
            added += 1
        ath["sessions"].sort(key=lambda x: x["date"], reverse=True)
    if added:
        print(f"[trackman_sync] filled {added} session(s) from bullpen reports "
              f"(provisional until master sheet catches up)", file=sys.stderr)


# ─── Entry point ─────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print('Usage: trackman_sync.py "<master.xlsx>" | master.csv [program_links.csv]',
              file=sys.stderr)
        sys.exit(2)

    src_path = sys.argv[1]
    if src_path.lower().endswith((".xlsx", ".xlsm")):
        sessions = list(load_master_xlsx(src_path))
        program_links = load_program_links_xlsx(src_path)
    else:
        sessions = list(load_master(src_path))
        program_links = load_program_links(sys.argv[2] if len(sys.argv) > 2 else None)

    portal = build_portal(sessions, program_links)

    # Sanity log to stderr
    print(
        f"[trackman_sync] {len(sessions)} sessions across "
        f"{len(portal['athletes'])} athletes",
        file=sys.stderr,
    )
    print(json.dumps(portal, indent=2))


if __name__ == "__main__":
    main()
