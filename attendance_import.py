#!/usr/bin/env python3
"""RPM Baseball Attendance workbook → attendance_manual.json

Frank's manually-tracked attendance: one sheet per week, columns
[Last Name, First Name, <date>, <date>, ...] with an "x" marking a day the
athlete was in the building. This is the fallback source for the profile
Consistency calendar — days with no ForceDecks/DynaMo data at all.

Usage:
    python3 attendance_import.py "/path/to/RPM Baseball Attendance.xlsx" > attendance_manual.json

Output: {"First Last": ["YYYY-MM-DD", ...], ...} sorted, deduped. Any non-empty
cell counts as attended (the sheet uses "x"). Commit the JSON; the portal
generator merges it on every build.
"""
import json
import sys
from datetime import datetime, date

import openpyxl


def main(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {}
    marks = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue
        dates = []
        for c in header[2:]:
            if isinstance(c, (datetime, date)):
                dates.append(c.strftime("%Y-%m-%d"))
            else:
                dates.append(None)
        for row in rows:
            if not row or (row[0] is None and row[1] is None):
                continue
            last = str(row[0] or "").strip()
            first = str(row[1] or "").strip()
            if not last and not first:
                continue
            name = f"{first} {last}".strip()
            for i, ds in enumerate(dates):
                if ds is None or 2 + i >= len(row):
                    continue
                v = row[2 + i]
                if v is None or str(v).strip() == "":
                    continue
                marks[str(v).strip().lower()] = marks.get(str(v).strip().lower(), 0) + 1
                out.setdefault(name, set()).add(ds)
    result = {n: sorted(ds) for n, ds in sorted(out.items())}
    total = sum(len(v) for v in result.values())
    print(f"[attendance] {len(result)} athletes, {total} attended days "
          f"across {len(wb.sheetnames)} week sheets; marks seen: {marks}",
          file=sys.stderr, flush=True)
    json.dump(result, sys.stdout, indent=0, separators=(",", ":"))


if __name__ == "__main__":
    main(sys.argv[1])
